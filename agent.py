import openpyxl
import numpy as np
import matplotlib.pyplot as plt

workbook = openpyxl.load_workbook('Data.xlsx')

dataTotal = 0

acessibilidadeTemp2020 = 0
data2020 = 0
matBasQT_Total2020 = 0
matBasFemQT_Total2020 = 0
matBasMascQT_Total2020 = 0
matBasNDQT_Total2020 = 0
matBasBrancaQT_Total2020 = 0
matBasPretaQT_total2020 = 0
matBasPardaQT_Total2020 = 0
matBasAmarelaQT_Total2020 = 0
matBasIndigenaQT_Total2020 = 0
compData2020 = 0
internetData2020 = 0
internetComum2020 = 0
matBas0_3QT_Total2020 = 0
matBas4_5QT_Total2020 = 0
matBas6_10QT_Total2020 = 0
matBas11_14QT_Total2020 = 0
matBas15_17QT_Total2020 = 0
matBas18PQT_Total2020 = 0

acessibilidadeTemp2021 = 0
data2021 = 0
matBasQT_Total2021 = 0
matBasFemQT_Total2021 = 0
matBasMascQT_Total2021 = 0
matBasNDQT_Total2021 = 0
matBasBrancaQT_Total2021 = 0
matBasPretaQT_total2021 = 0
matBasPardaQT_Total2021 = 0
matBasAmarelaQT_Total2021 = 0
matBasIndigenaQT_Total2021 = 0
compData2021 = 0
internetData2021 = 0
internetComum2021 = 0
matBas0_3QT_Total2021 = 0
matBas4_5QT_Total2021 = 0
matBas6_10QT_Total2021 = 0
matBas11_14QT_Total2021 = 0
matBas15_17QT_Total2021 = 0
matBas18PQT_Total2021 = 0


sheet = workbook['Data_1']

for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
    year_data = row[0]
    acessibility1_data = row[3]
    acessibility2_data = row[4]
    acessibility3_data = row[5]
    acessibility4_data = row[6]
    acessibility5_data = row[7]
    acessibility6_data = row[8]
    acessibility7_data = row[9]
    acessibility8_data = row[10]
    noAcessibility_data = row[11]
    desktopAluno_data = row[12]
    tabletAluno_data = row[13]
    internet_data = row[14]
    internetAluno_data = row[15]
    internetAprend_data = row[16]
    internetComun_data = row[17]
    InternetCompAcess_data = row[18]
    internetDispPesso_data = row[19]
    matBasQT_data = row[20]
    matBasFemQT_data = row[21]
    matBasMascQT_data = row[22]
    matBasNDQT_data = row[23]
    matBasBrancaQT_data = row[24]
    matBasPretaQT_data = row[25]
    matBasPardaQT_data = row[26]
    matBasAmarelaQT_data = row[27]
    matBasIndigenaQT_data = row[28]
    matBas0_3QT_data = row[29]
    matBas4_5QT_data = row[30]
    matBas6_10QT_data = row[31]
    matBas11_14QT_data = row[32]
    matBas15_17QT_data = row[33]
    matBas18PQT_data = row[34]
    matBasDQT_data = row[35]
    matBasNQT_data = row[36]
    matBasEAD_data = row[37]

    dataTotal += 1

    if year_data == 2020:
        data2020 += 1
    if year_data == 2020 and matBasQT_data is not None:
        matBasQT_Total2020 += matBasQT_data
    if year_data == 2020 and matBasFemQT_data is not None:
        matBasFemQT_Total2020 += matBasFemQT_data
    if year_data == 2020 and matBasMascQT_data is not None:
        matBasMascQT_Total2020 += matBasMascQT_data
    if year_data == 2020 and matBasBrancaQT_data is not None:
        matBasBrancaQT_Total2020 += matBasBrancaQT_data
    if year_data == 2020 and matBasPretaQT_data is not None:
        matBasPretaQT_total2020 += matBasPretaQT_data
    if year_data == 2020 and matBasPardaQT_data is not None:
        matBasPardaQT_Total2020 += matBasPardaQT_data
    if year_data == 2020 and matBasAmarelaQT_data is not None:
        matBasAmarelaQT_Total2020 += matBasAmarelaQT_data
    if year_data == 2020 and matBasIndigenaQT_data is not None:
        matBasIndigenaQT_Total2020 += matBasIndigenaQT_data
    if year_data == 2020 and noAcessibility_data == 1:
        acessibilidadeTemp2020 += 1
        acessibilidade2020 = data2020 - acessibilidadeTemp2020
    if year_data == 2020 and (tabletAluno_data == 1 or desktopAluno_data == 1):
        compData2020 += 1
    if year_data == 2020 and (internet_data == 1 or internetAluno_data == 1 or internetAprend_data == 1 or internetComun_data == 1 or InternetCompAcess_data == 1):
        internetData2020 += 1
    if year_data == 2020 and internetDispPesso_data == 1:
        internetComum2020 += 1
    if year_data == 2020 and matBas0_3QT_data is not None:
        matBas0_3QT_Total2020 += matBas0_3QT_data
    if year_data == 2020 and matBas4_5QT_data is not None:
        matBas4_5QT_Total2020 += matBas4_5QT_data
    if year_data == 2020 and matBas6_10QT_data is not None:
        matBas6_10QT_Total2020 += matBas6_10QT_data
    if year_data == 2020 and matBas11_14QT_data is not None:
        matBas11_14QT_Total2020 += matBas11_14QT_data
    if year_data == 2020 and matBas15_17QT_data is not None:
        matBas15_17QT_Total2020 += matBas15_17QT_data
    if year_data == 2020 and matBas18PQT_data is not None:
        matBas18PQT_Total2020 += matBas18PQT_data
        
    

    if year_data == 2021:
        data2021 += 1
    if year_data == 2021 and matBasQT_data is not None:
        matBasQT_Total2021 += matBasQT_data
    if year_data == 2021 and matBasFemQT_data is not None:
        matBasFemQT_Total2021 += matBasFemQT_data
    if year_data == 2021 and matBasMascQT_data is not None:
        matBasMascQT_Total2021 += matBasMascQT_data
    if year_data == 2021 and matBasBrancaQT_data is not None:
        matBasBrancaQT_Total2021 += matBasBrancaQT_data
    if year_data == 2021 and matBasPretaQT_data is not None:
        matBasPretaQT_total2021 += matBasPretaQT_data
    if year_data == 2021 and matBasPardaQT_data is not None:
        matBasPardaQT_Total2021 += matBasPardaQT_data
    if year_data == 2021 and matBasAmarelaQT_data is not None:
        matBasAmarelaQT_Total2021 += matBasAmarelaQT_data
    if year_data == 2021 and matBasIndigenaQT_data is not None:
        matBasIndigenaQT_Total2021 += matBasIndigenaQT_data
    if year_data == 2021 and noAcessibility_data == 1:
        acessibilidadeTemp2021 += 1
        acessibilidade2021 = data2021 - acessibilidadeTemp2021
    if year_data == 2021 and (tabletAluno_data == 1 or desktopAluno_data == 1):
        compData2021 += 1
    if year_data == 2021 and (internet_data == 1 or internetAluno_data == 1 or internetAprend_data == 1 or internetComun_data == 1 or InternetCompAcess_data == 1):
        internetData2021 += 1
    if year_data == 2021 and internetDispPesso_data == 1:
        internetComum2021 += 1
    if year_data == 2021 and matBas0_3QT_data is not None:
        matBas0_3QT_Total2021 += matBas0_3QT_data
    if year_data == 2021 and matBas4_5QT_data is not None:
        matBas4_5QT_Total2021 += matBas4_5QT_data
    if year_data == 2021 and matBas6_10QT_data is not None:
        matBas6_10QT_Total2021 += matBas6_10QT_data
    if year_data == 2021 and matBas11_14QT_data is not None:
        matBas11_14QT_Total2021 += matBas11_14QT_data
    if year_data == 2021 and matBas15_17QT_data is not None:
        matBas15_17QT_Total2021 += matBas15_17QT_data
    if year_data == 2021 and matBas18PQT_data is not None:
        matBas18PQT_Total2021 += matBas18PQT_data
    

print(f'Foram avalidas {dataTotal} instituições de ensino entre 2020 e 2021, dentre todas a quantidade de alunos matriculados no ensino basico na bahia em 2020 foi {matBasQT_Total2020} e em 2021 foi {matBasQT_Total2021}.')
print(f'===========================================================================================================================')
print(f'2020: Foram encontrados {matBasQT_Total2020} alunos matriculados em 2020.')
print(f'2020: Deste {matBasQT_Total2020} alunos: {matBasFemQT_Total2020} são mulheres e {matBasMascQT_Total2020} são homens.')
print(f'2020: Além disso: {matBasBrancaQT_Total2020} são brancos, {matBasPretaQT_total2020} são negros, {matBasPardaQT_Total2020} são pardos, {matBasAmarelaQT_Total2020} são amarelos e {matBasIndigenaQT_Total2020} são indigenas.')
print(f'2020: Foram analisadas {data2020} instituições e entre estas apenas {acessibilidade2020} possui alguma forma de acessibilidade.')
print(f'2020: Destas apenas {compData2020} possui computadores ou tablets disponiveis para os alunos')
print(f'2020: Destas apenas {internetData2020} possui acesso a internet para os computadores internos')
print(f'2020: Destas apenas {internetComum2020} possui acesso a internet disponivel para os alunos')
print(f'2020: A quantidade de alunos no ensino medio por faixa etaria é: 0-3: {matBas0_3QT_Total2020}; 4-5: {matBas4_5QT_Total2020}; 6-10: {matBas6_10QT_Total2020}; 11-14: {matBas11_14QT_Total2020}; 15-17: {matBas15_17QT_Total2020}; 18+: {matBas18PQT_Total2020}')
print(f'===========================================================================================================================')
print(f'2021: Foram encontrados {matBasQT_Total2021} alunos matriculados em 2021.')
print(f'2021: Deste {matBasQT_Total2021} alunos: {matBasFemQT_Total2021} são mulheres e {matBasMascQT_Total2021} são homens.')
print(f'2021: Além disso: {matBasBrancaQT_Total2021} são brancos, {matBasPretaQT_total2021} são negros, {matBasPardaQT_Total2021} são pardos, {matBasAmarelaQT_Total2021} são amarelos e {matBasIndigenaQT_Total2021} são indigenas.')
print(f'2021: Foram analisadas {data2021} instituições e entre estas apenas {acessibilidade2021} possui alguma forma de acessibilidade.')
print(f'2021: Destas apenas {compData2021} possui computadores ou tablets disponiveis para os alunos')
print(f'2021: Destas apenas {internetData2021} possui acesso a internet para os computadores internos')
print(f'2021: Destas apenas {internetComum2021} possui acesso a internet disponivel para os alunos')
print(f'2021: A quantidade de alunos no ensino medio por faixa etaria é: 0-3: {matBas0_3QT_Total2021}; 4-5: {matBas4_5QT_Total2021}; 6-10: {matBas6_10QT_Total2021}; 11-14: {matBas11_14QT_Total2021}; 15-17: {matBas15_17QT_Total2021}; 18+: {matBas18PQT_Total2021}')

matBasFemPercentage2020 = ((matBasFemQT_Total2020 * 100) / matBasQT_Total2020)
matBasBrancaPercentage2020 = ((matBasBrancaQT_Total2020 * 100) / matBasQT_Total2020)
matBasPretaPercentage2020 = ((matBasPretaQT_total2020 * 100) / matBasQT_Total2020)
matBasPardaPercentage2020 = ((matBasPardaQT_Total2020 * 100) / matBasQT_Total2020)
matBasAmarelaPercentage2020 = ((matBasAmarelaQT_Total2020 * 100) / matBasQT_Total2020)
matBasIndigenaPercentage2020 = ((matBasIndigenaQT_Total2020 * 100) / matBasQT_Total2020)
acessibilidadePercentage2020 = ((acessibilidade2020 * 100) / data2020)
computadoresPercentage2020 = ((compData2020 * 100) / data2020)
internetPercentage2020 = ((internetData2020 * 100) / data2020)
InternetComumPercentage2020 = ((internetComum2020 * 100) / data2020)
matBas0_3Percentage2020 = ((matBas0_3QT_Total2020 * 100) / matBasQT_Total2020)
matBas4_5Percentage2020 = ((matBas4_5QT_Total2020 * 100) / matBasQT_Total2020)
matBas6_10Percentage2020 = ((matBas6_10QT_Total2020 * 100) / matBasQT_Total2020)
matBas11_14Percentage2020 = ((matBas11_14QT_Total2020 * 100) / matBasQT_Total2020)
matBas15_17Percentage2020 = ((matBas15_17QT_Total2020 * 100) / matBasQT_Total2020)
matbas18PPercentage2020 = ((matBas18PQT_Total2020 * 100) / matBasQT_Total2020)

matBasFemPercentage2021 = ((matBasFemQT_Total2021 * 100) / matBasQT_Total2021)
matBasBrancaPercentage2021 = ((matBasBrancaQT_Total2021 * 100) / matBasQT_Total2021)
matBasPretaPercentage2021 = ((matBasPretaQT_total2021 * 100) / matBasQT_Total2021)
matBasPardaPercentage2021 = ((matBasPardaQT_Total2021 * 100) / matBasQT_Total2021)
matBasAmarelaPercentage2021 = ((matBasAmarelaQT_Total2021 * 100) / matBasQT_Total2021)
matBasIndigenaPercentage2021 = ((matBasIndigenaQT_Total2021 * 100) / matBasQT_Total2021)
acessibilidadePercentage2021 = ((acessibilidade2021 * 100) / data2021)
computadoresPercentage2021 = ((compData2021 * 100) / data2021)
internetPercentage2021 = ((internetData2021 * 100) / data2021)
InternetComumPercentage2021 = ((internetComum2021 * 100) / data2021)
matBas0_3Percentage2021 = ((matBas0_3QT_Total2021 * 100) / matBasQT_Total2021)
matBas4_5Percentage2021 = ((matBas4_5QT_Total2021 * 100) / matBasQT_Total2021)
matBas6_10Percentage2021 = ((matBas6_10QT_Total2021 * 100) / matBasQT_Total2021)
matBas11_14Percentage2021 = ((matBas11_14QT_Total2021 * 100) / matBasQT_Total2021)
matBas15_17Percentage2021 = ((matBas15_17QT_Total2021 * 100) / matBasQT_Total2021)
matbas18PPercentage2021 = ((matBas18PQT_Total2021 * 100) / matBasQT_Total2021)

dataGender2020 = [matBasFemPercentage2020, 100-matBasFemPercentage2020]
dataRace2020 = [matBasBrancaPercentage2020, matBasPretaPercentage2020, matBasPardaPercentage2020, matBasAmarelaPercentage2020, matBasIndigenaPercentage2020]
dataAcessibilidade2020 = [acessibilidadePercentage2020, 100-acessibilidadePercentage2020]
dataComputadores2020 = [computadoresPercentage2020, 100-computadoresPercentage2020]
dataInternet2020 = [internetPercentage2020, 100-internetPercentage2020]
dataInternetComum2020 = [InternetComumPercentage2020, 100-InternetComumPercentage2020]
dataAge2020 = [matBas0_3Percentage2020, matBas4_5Percentage2020, matBas6_10Percentage2020, matBas11_14Percentage2020, matBas15_17Percentage2020, matbas18PPercentage2020]

dataGender2021 = [matBasFemPercentage2021, 100-matBasFemPercentage2021]
dataRace2021 = [matBasBrancaPercentage2021, matBasPretaPercentage2021, matBasPardaPercentage2021, matBasAmarelaPercentage2021, matBasIndigenaPercentage2021]
dataAcessibilidade2021 = [acessibilidadePercentage2021, 100-acessibilidadePercentage2021]
dataComputadores2021 = [computadoresPercentage2021, 100-computadoresPercentage2021]
dataInternet2021 = [internetPercentage2021, 100-internetPercentage2021]
dataInternetComum2021 = [InternetComumPercentage2021, 100-InternetComumPercentage2021]
dataAge2021 = [matBas0_3Percentage2021, matBas4_5Percentage2021, matBas6_10Percentage2021, matBas11_14Percentage2021, matBas15_17Percentage2021, matbas18PPercentage2021]

fig, axs = plt.subplots(2, 4, figsize=(20, 8))

axs[0, 0].pie(dataGender2020, autopct='%1.1f%%', startangle=90 )
axs[0, 0].set_title('Dados de alunos entre \nhomens e mulheres em 2020.')
axs[0, 0].legend(['Mulheres', 'Homens'],bbox_to_anchor=(1, 0.5), loc="center left")

axs[0, 1].pie(dataComputadores2020,labels=['Com acesso a computadoes', 'Sem acesso a computadores'], autopct='%1.1f%%', startangle=90)
axs[0, 1].set_title('Instituiçoes com acesso a computadores')

axs[0, 2].pie(dataAcessibilidade2020,labels=['Com acessibilidade', 'Sem acessibilidade'], autopct='%1.1f%%', startangle=90)
axs[0, 2].set_title('Instituições com acessibilidade em 2020')

axs[1, 0].pie(dataRace2020, autopct='%1.1f%%', startangle=90)
axs[1, 0].set_title('Dados de alunos \npor cor em 2020.')
axs[1, 0].legend(['Brancos', 'Negros', 'Pardos', 'Amarelo', 'Indigena'],bbox_to_anchor=(1, 0.5), loc="center left")

axs[1, 1].pie(dataInternet2020, labels=['Com acesso a internet', 'Sem acesso a internet'], autopct='%1.1f%%', startangle=45)
axs[1, 1].set_title('Dados de acesso a internet em 2020.')

axs[1, 2].pie(dataInternetComum2020, labels=['Com acesso a internet', 'Sem acesso a internet'], autopct='%1.1f%%', startangle=90)
axs[1, 2].set_title('Dados de acesso comum a internet em 2020.')

axs[0, 3].pie(dataAge2020, autopct='%1.1f%%', startangle=90)
axs[0, 3].set_title('Dados de idade \nentre 0 a 18 anos em 2020.')
axs[0, 3].legend(['0-3', '4-5', '6-10', '11-14', '15-17', '18+'],loc='upper center', bbox_to_anchor=(0.5, -0.2), fancybox=True, shadow=True, ncol=3)

plt.tight_layout()
plt.show()

fig2, axs2 = plt.subplots(2, 4, figsize=(20, 8))

axs2[0, 0].pie(dataGender2021, autopct='%1.1f%%', startangle=90 )
axs2[0, 0].set_title('Dados de alunos entre \nhomens e mulheres em 2021.')
axs2[0, 0].legend(['Mulheres', 'Homens'],bbox_to_anchor=(1, 0.5), loc="center left")

axs2[0, 1].pie(dataComputadores2021,labels=['Com acesso a computadoes', 'Sem acesso a computadores'], autopct='%1.1f%%', startangle=90)
axs2[0, 1].set_title('Instituiçoes com acesso a computadores')

axs2[0, 2].pie(dataAcessibilidade2021,labels=['Com acessibilidade', 'Sem acessibilidade'], autopct='%1.1f%%', startangle=90)
axs2[0, 2].set_title('Instituições com acessibilidade em 2021')

axs2[1, 0].pie(dataRace2021, autopct='%1.1f%%', startangle=90)
axs2[1, 0].set_title('Dados de alunos \npor cor em 2021.')
axs2[1, 0].legend(['Brancos', 'Negros', 'Pardos', 'Amarelo', 'Indigena'],bbox_to_anchor=(1, 0.5), loc="center left")

axs2[1, 1].pie(dataInternet2021, labels=['Com acesso a internet', 'Sem acesso a internet'], autopct='%1.1f%%', startangle=90)
axs2[1, 1].set_title('Dados de acesso a internet em 2021.')

axs2[1, 2].pie(dataInternetComum2021, labels=['Com acesso a internet', 'Sem acesso a internet'], autopct='%1.1f%%', startangle=0)
axs2[1, 2].set_title('Dados de acesso comum a internet em 2021.')

axs2[0, 3].pie(dataAge2021, autopct='%1.1f%%', startangle=90)
axs2[0, 3].set_title('Dados de idade \nentre 0 a 18 anos em 2021.')
axs2[0, 3].legend(['0-3', '4-5', '6-10', '11-14', '15-17', '18+'],loc='upper center', bbox_to_anchor=(0.5, -0.2), fancybox=True, shadow=True, ncol=3)

plt.tight_layout()
plt.show()

fig3, axs3 = plt.subplots(2, 4, figsize=(20, 8))

categories_gender = ['Mulheres', 'Homens']
data_gender_2020 = [matBasFemPercentage2020, 100 - matBasFemPercentage2020]
data_gender_2021 = [matBasFemPercentage2021, 100 - matBasFemPercentage2021]

bar_width = 0.35
bar_positions1 = np.arange(len(categories_gender))
bar_positions2 = [pos + bar_width for pos in bar_positions1]

axs3[0, 0].bar(bar_positions1, data_gender_2020, width=bar_width, label='2020')
axs3[0, 0].bar(bar_positions2, data_gender_2021, width=bar_width, label='2021')

axs3[0, 0].set_xticks([pos + bar_width / 2 for pos in bar_positions1])
axs3[0, 0].set_xticklabels(categories_gender)
axs3[0, 0].set_title('Dados de alunos entre homens e mulheres')
axs3[0, 0].legend(loc='upper center', bbox_to_anchor=(0.5, -0.2), fancybox=True, shadow=True, ncol=3)

categories_race = ['Brancos', 'Negros', 'Pardos', 'Amarelo', 'Indigena']
data_race_2020 = [matBasBrancaPercentage2020, matBasPretaPercentage2020, matBasPardaPercentage2020, matBasAmarelaPercentage2020, matBasIndigenaPercentage2020]
data_race_2021 = [matBasBrancaPercentage2021, matBasPretaPercentage2021, matBasPardaPercentage2021, matBasAmarelaPercentage2021, matBasIndigenaPercentage2021]

bar_positions1_race = np.arange(len(categories_race))
bar_positions2_race = [pos + bar_width for pos in bar_positions1_race]

axs3[0, 1].bar(bar_positions1_race, data_race_2020, width=bar_width, label='2020')
axs3[0, 1].bar(bar_positions2_race, data_race_2021, width=bar_width, label='2021')

axs3[0, 1].set_xticks([pos + bar_width / 2 for pos in bar_positions1_race])
axs3[0, 1].set_xticklabels(categories_race)
axs3[0, 1].set_title('Dados de alunos por cor')
axs3[0, 1].legend(loc='upper center', bbox_to_anchor=(0.5, -0.2), fancybox=True, shadow=True, ncol=3)

categories_acessibilidade = ['Com acessibilidade', 'Sem acessibilidade']
data_acessibilidade_2020 = [acessibilidadePercentage2020, 100 - acessibilidadePercentage2020]
data_acessibilidade_2021 = [acessibilidadePercentage2021, 100 - acessibilidadePercentage2021]

bar_positions1_acessibilidade = np.arange(len(categories_acessibilidade))
bar_positions2_acessibilidade = [pos + bar_width for pos in bar_positions1_acessibilidade]

axs3[0, 2].bar(bar_positions1_acessibilidade, data_acessibilidade_2020, width=bar_width, label='2020')
axs3[0, 2].bar(bar_positions2_acessibilidade, data_acessibilidade_2021, width=bar_width, label='2021')

axs3[0, 2].set_xticks([pos + bar_width / 2 for pos in bar_positions1_acessibilidade])
axs3[0, 2].set_xticklabels(categories_acessibilidade)
axs3[0, 2].set_title('Instituições com acessibilidade')
axs3[0, 2].legend(loc='upper center', bbox_to_anchor=(0.5, -0.2), fancybox=True, shadow=True, ncol=3)

categories_computadores = ['Com acesso', 'Sem acesso']
data_computadores_2020 = [computadoresPercentage2020, 100 - computadoresPercentage2020]
data_computadores_2021 = [computadoresPercentage2021, 100 - computadoresPercentage2021]

bar_positions1_computadores = np.arange(len(categories_computadores))
bar_positions2_computadores = [pos + bar_width for pos in bar_positions1_computadores]

axs3[1, 0].bar(bar_positions1_computadores, data_computadores_2020, width=bar_width, label='2020')
axs3[1, 0].bar(bar_positions2_computadores, data_computadores_2021, width=bar_width, label='2021')

axs3[1, 0].set_xticks([pos + bar_width / 2 for pos in bar_positions1_computadores])
axs3[1, 0].set_xticklabels(categories_computadores)
axs3[1, 0].set_title('Instituições com acesso a computadores')
axs3[1, 0].legend(loc='upper center', bbox_to_anchor=(0.5, -0.2), fancybox=True, shadow=True, ncol=3)

categories_internet = ['Com acesso à internet', 'Sem acesso à internet']
data_internet_2020 = [internetPercentage2020, 100 - internetPercentage2020]
data_internet_2021 = [internetPercentage2021, 100 - internetPercentage2021]

bar_positions1_internet = np.arange(len(categories_internet))
bar_positions2_internet = [pos + bar_width for pos in bar_positions1_internet]

axs3[1, 1].bar(bar_positions1_internet, data_internet_2020, width=bar_width, label='2020')
axs3[1, 1].bar(bar_positions2_internet, data_internet_2021, width=bar_width, label='2021')

axs3[1, 1].set_xticks([pos + bar_width / 2 for pos in bar_positions1_internet])
axs3[1, 1].set_xticklabels(categories_internet)
axs3[1, 1].set_title('Dados de acesso à internet')
axs3[1, 1].legend(loc='upper center', bbox_to_anchor=(0.5, -0.2), fancybox=True, shadow=True, ncol=3)

categories_internet_comum = ['Com acesso à internet', 'Sem acesso à internet']
data_internet_comum_2020 = [InternetComumPercentage2020, 100 - InternetComumPercentage2020]
data_internet_comum_2021 = [InternetComumPercentage2021, 100 - InternetComumPercentage2021]

bar_positions1_internet_comum = np.arange(len(categories_internet_comum))
bar_positions2_internet_comum = [pos + bar_width for pos in bar_positions1_internet_comum]

axs3[1, 2].bar(bar_positions1_internet_comum, data_internet_comum_2020, width=bar_width, label='2020')
axs3[1, 2].bar(bar_positions2_internet_comum, data_internet_comum_2021, width=bar_width, label='2021')

axs3[1, 2].set_xticks([pos + bar_width / 2 for pos in bar_positions1_internet_comum])
axs3[1, 2].set_xticklabels(categories_internet_comum)
axs3[1, 2].set_title('Dados de acesso comum à internet')
axs3[1, 2].legend(loc='upper center', bbox_to_anchor=(0.5, -0.2), fancybox=True, shadow=True, ncol=3)

categories_idade = ['0-3', '4-5', '6-10', '11-14', '15-17', '18+']
data_idade_2020 = [matBas0_3Percentage2020, matBas4_5Percentage2020, matBas6_10Percentage2020, matBas11_14Percentage2020, matBas15_17Percentage2020, matbas18PPercentage2020]
data_idade_2021 = [matBas0_3Percentage2021, matBas4_5Percentage2021, matBas6_10Percentage2021, matBas11_14Percentage2021, matBas15_17Percentage2021, matbas18PPercentage2021]

bar_positions1_idade = np.arange(len(categories_idade))
bar_positions2_idade = [pos + bar_width for pos in bar_positions1_idade]

axs3[1, 3].bar(bar_positions1_idade, data_idade_2020, width=bar_width, label='2020')
axs3[1, 3].bar(bar_positions2_idade, data_idade_2021, width=bar_width, label='2021')

axs3[1, 3].set_xticks([pos + bar_width / 2 for pos in bar_positions1_idade])
axs3[1, 3].set_xticklabels(categories_idade)
axs3[1, 3].set_title('Dados de idade entre 0 a 18 anos')
axs3[1, 3].legend(loc='upper center', bbox_to_anchor=(0.5, -0.2), fancybox=True, shadow=True, ncol=3)

plt.tight_layout()
plt.show()